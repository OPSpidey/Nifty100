import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.append(str(Path(__file__).resolve().parents[2]))

from src.screener.engine import load_config, run_all_presets

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_OUTPUT = PROJECT_ROOT / "output" / "screener_output.xlsx"


KPI_COLUMNS = [
    "company_id",
    "year",
    "broad_sector",
    "composite_quality_score",
    "return_on_equity_pct",
    "roce_pct",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",
    "debt_to_equity",
    "interest_coverage_for_filter",
    "asset_turnover",
    "free_cash_flow_cr",
    "fcf_cagr_5yr",
    "cfo_pat_ratio",
    "revenue_cagr_5yr",
    "revenue_cagr_3yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",
    "pe_ratio",
    "pb_ratio",
    "dividend_yield_pct",
    "dividend_payout_ratio_pct",
    "sales",
    "net_profit",
    "market_cap_crore",
]


THRESHOLD_COLUMNS = {
    "roe_min": "return_on_equity_pct",
    "debt_to_equity_max": "debt_to_equity",
    "debt_to_equity_eq": "debt_to_equity",
    "free_cash_flow_min": "free_cash_flow_cr",
    "revenue_cagr_5yr_min": "revenue_cagr_5yr",
    "revenue_cagr_3yr_min": "revenue_cagr_3yr",
    "pat_cagr_5yr_min": "pat_cagr_5yr",
    "pe_max": "pe_ratio",
    "pb_max": "pb_ratio",
    "dividend_yield_min": "dividend_yield_pct",
    "dividend_payout_max": "dividend_payout_ratio_pct",
    "sales_min": "sales",
}


def _passes(value, filter_name, threshold):
    if value is None:
        return False

    if filter_name.endswith("_min"):
        return value >= threshold
    if filter_name.endswith("_max"):
        return value <= threshold
    if filter_name.endswith("_eq"):
        return value == threshold

    return True


def _style_sheet(sheet, df, preset_filters):
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    green_fill = PatternFill("solid", fgColor="D1FAE5")
    red_fill = PatternFill("solid", fgColor="FEE2E2")
    title_fill = PatternFill("solid", fgColor="E5E7EB")

    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    column_lookup = {cell.value: cell.column for cell in sheet[1]}

    for filter_name, threshold in preset_filters.items():
        column = THRESHOLD_COLUMNS.get(filter_name)
        if column not in column_lookup:
            continue

        column_index = column_lookup[column]

        for row_index in range(2, len(df) + 2):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.fill = green_fill if _passes(cell.value, filter_name, threshold) else red_fill

    for column_index, column_name in enumerate(KPI_COLUMNS, start=1):
        letter = get_column_letter(column_index)
        width = max(12, min(24, len(column_name) + 2))
        sheet.column_dimensions[letter].width = width

        if column_name not in {"company_id", "year", "broad_sector"}:
            for cell in sheet[letter][1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00"

    if not df.empty:
        score_col = column_lookup.get("composite_quality_score")
        if score_col:
            for row_index in range(2, len(df) + 2):
                sheet.cell(row=row_index, column=score_col).fill = title_fill


def export_screener_output(output_path=DEFAULT_OUTPUT):
    config = load_config()
    presets = run_all_presets()

    workbook = Workbook()
    workbook.remove(workbook.active)

    for preset_name, df in presets.items():
        sheet = workbook.create_sheet(preset_name[:31])
        out = df.reindex(columns=KPI_COLUMNS)
        sheet.append(KPI_COLUMNS)

        for row in out.itertuples(index=False):
            sheet.append(list(row))

        _style_sheet(sheet, out, config["presets"][preset_name])
        sheet.auto_filter.ref = sheet.dimensions

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    return output_path


if __name__ == "__main__":
    print(export_screener_output())
