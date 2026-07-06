from pathlib import Path
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.append(str(Path(__file__).resolve().parents[2]))

from src.analytics.peer import PEER_METRICS, REPORT_METRICS, peer_report_frame


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_OUTPUT = PROJECT_ROOT / "output" / "peer_comparison.xlsx"


BASE_COLUMNS = [
    "company_id",
    "company_name",
    "year",
    "is_benchmark",
]

PERCENTILE_COLUMNS = [f"{metric} Percentile" for metric in PEER_METRICS]

OUTPUT_COLUMNS = BASE_COLUMNS + REPORT_METRICS + PERCENTILE_COLUMNS


def _style_sheet(sheet, row_count):
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    benchmark_fill = PatternFill("solid", fgColor="FDE68A")
    green_fill = PatternFill("solid", fgColor="DCFCE7")
    yellow_fill = PatternFill("solid", fgColor="FEF9C3")
    red_fill = PatternFill("solid", fgColor="FEE2E2")
    median_fill = PatternFill("solid", fgColor="E5E7EB")

    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    header_lookup = {cell.value: cell.column for cell in sheet[1]}
    benchmark_col = header_lookup.get("is_benchmark")

    for row_index in range(2, row_count + 2):
        if benchmark_col and sheet.cell(row=row_index, column=benchmark_col).value == 1:
            for cell in sheet[row_index]:
                cell.fill = benchmark_fill

    for column in PERCENTILE_COLUMNS:
        column_index = header_lookup.get(column)
        if not column_index:
            continue

        for row_index in range(2, row_count + 2):
            cell = sheet.cell(row=row_index, column=column_index)
            if cell.value is None:
                continue
            cell.number_format = "0.0%"
            if cell.value >= 0.75:
                cell.fill = green_fill
            elif cell.value <= 0.25:
                cell.fill = red_fill
            else:
                cell.fill = yellow_fill

    median_row = row_count + 2
    for cell in sheet[median_row]:
        cell.fill = median_fill
        cell.font = Font(bold=True)

    for column_index, column_name in enumerate(OUTPUT_COLUMNS, start=1):
        letter = get_column_letter(column_index)
        sheet.column_dimensions[letter].width = max(12, min(26, len(column_name) + 2))

        if column_name not in {"company_id", "company_name", "year"}:
            for cell in sheet[letter][1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00"

    sheet.auto_filter.ref = sheet.dimensions


def export_peer_comparison(output_path=DEFAULT_OUTPUT):
    report = peer_report_frame()
    workbook = Workbook()
    workbook.remove(workbook.active)

    for peer_group_name in sorted(report["peer_group_name"].dropna().unique()):
        group = report[report["peer_group_name"].eq(peer_group_name)].copy()
        group = group.sort_values(
            ["is_benchmark", "composite_quality_score"],
            ascending=[False, False],
        )
        sheet = workbook.create_sheet(peer_group_name[:31])
        out = group.reindex(columns=OUTPUT_COLUMNS)
        sheet.append(OUTPUT_COLUMNS)

        for row in out.itertuples(index=False):
            sheet.append(list(row))

        median_row = ["Median", "", "", ""]
        for column in REPORT_METRICS:
            median_row.append(group[column].median(skipna=True))
        for column in PERCENTILE_COLUMNS:
            median_row.append(group[column].median(skipna=True))
        sheet.append(median_row)

        _style_sheet(sheet, len(out))

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    return output_path


if __name__ == "__main__":
    print(export_peer_comparison())
